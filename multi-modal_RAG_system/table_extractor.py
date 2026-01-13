"""
Table Extractor

Extracts and processes tables from:
- CSV/TSV files
- Excel files
- HTML tables
- Markdown tables
- Images of tables (via vision)
"""

from dataclasses import dataclass
from typing import Optional, Any, Union
from pathlib import Path
import io
import re
import json

from .content_types import (
    TableContent,
    ContentType,
    generate_content_id,
)


@dataclass
class TableExtractionConfig:
    """Configuration for table extraction."""
    
    # Chunking
    max_rows_per_chunk: int = 50
    include_headers_in_chunks: bool = True
    
    # Processing
    infer_types: bool = True
    clean_whitespace: bool = True
    
    # Output
    generate_summary: bool = True


class TableExtractor:
    """
    Extracts tables from various formats.
    
    Supports:
    - CSV/TSV files
    - Excel files (.xlsx, .xls)
    - HTML tables
    - Markdown tables
    - JSON arrays
    """
    
    def __init__(self, config: Optional[TableExtractionConfig] = None):
        self.config = config or TableExtractionConfig()
    
    def extract_file(self, file_path: str) -> list[TableContent]:
        """Extract tables from a file based on extension."""
        path = Path(file_path)
        ext = path.suffix.lower()
        
        if ext == ".csv":
            return [self.extract_csv(file_path)]
        elif ext == ".tsv":
            return [self.extract_csv(file_path, delimiter="\t")]
        elif ext in (".xlsx", ".xls"):
            return self.extract_excel(file_path)
        elif ext == ".html":
            return self.extract_html_file(file_path)
        elif ext == ".md":
            return self.extract_markdown_file(file_path)
        elif ext == ".json":
            return [self.extract_json(file_path)]
        else:
            raise ValueError(f"Unsupported file type: {ext}")
    
    def extract_csv(
        self,
        file_path: str,
        delimiter: str = ",",
        encoding: str = "utf-8",
    ) -> TableContent:
        """Extract table from CSV/TSV file."""
        import csv
        
        rows = []
        headers = []
        
        with open(file_path, "r", encoding=encoding, errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            
            for i, row in enumerate(reader):
                if i == 0:
                    headers = [self._clean_cell(h) for h in row]
                else:
                    cleaned_row = [self._clean_cell(cell) for cell in row]
                    rows.append(cleaned_row)
        
        return self._create_table_content(
            headers=headers,
            rows=rows,
            source_file=file_path,
        )
    
    def extract_excel(self, file_path: str) -> list[TableContent]:
        """Extract tables from Excel file (all sheets)."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Install openpyxl for Excel support: pip install openpyxl")
        
        tables = []
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Convert to list of lists
            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None for cell in row):
                    rows.append([self._clean_cell(cell) for cell in row])
            
            if not rows:
                continue
            
            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []
            
            table = self._create_table_content(
                headers=headers,
                rows=data_rows,
                source_file=file_path,
                caption=sheet_name,
            )
            table.metadata["sheet_name"] = sheet_name
            
            tables.append(table)
        
        wb.close()
        return tables
    
    def extract_html(self, html_content: str, source_file: str = "html") -> list[TableContent]:
        """Extract tables from HTML content."""
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            raise ImportError("Install beautifulsoup4 for HTML support: pip install beautifulsoup4")
        
        soup = BeautifulSoup(html_content, "html.parser")
        tables = []
        
        for i, table_elem in enumerate(soup.find_all("table")):
            headers = []
            rows = []
            
            # Extract headers
            header_row = table_elem.find("thead")
            if header_row:
                headers = [
                    self._clean_cell(th.get_text())
                    for th in header_row.find_all(["th", "td"])
                ]
            
            # Extract rows
            body = table_elem.find("tbody") or table_elem
            for tr in body.find_all("tr"):
                cells = tr.find_all(["td", "th"])
                
                if not headers and all(c.name == "th" for c in cells):
                    headers = [self._clean_cell(c.get_text()) for c in cells]
                else:
                    row = [self._clean_cell(c.get_text()) for c in cells]
                    if row:
                        rows.append(row)
            
            # Get caption if available
            caption = None
            caption_elem = table_elem.find("caption")
            if caption_elem:
                caption = caption_elem.get_text().strip()
            
            if headers or rows:
                table = self._create_table_content(
                    headers=headers,
                    rows=rows,
                    source_file=source_file,
                    caption=caption,
                )
                table.html = str(table_elem)
                tables.append(table)
        
        return tables
    
    def extract_html_file(self, file_path: str) -> list[TableContent]:
        """Extract tables from an HTML file."""
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        return self.extract_html(content, file_path)
    
    def extract_markdown(self, md_content: str, source_file: str = "markdown") -> list[TableContent]:
        """Extract tables from Markdown content."""
        tables = []
        
        # Regex to match markdown tables
        table_pattern = re.compile(
            r'(?:^|\n)((?:\|[^\n]+\|(?:\n|$))+)',
            re.MULTILINE
        )
        
        for i, match in enumerate(table_pattern.finditer(md_content)):
            table_text = match.group(1).strip()
            lines = table_text.split("\n")
            
            if len(lines) < 2:
                continue
            
            headers = []
            rows = []
            
            for j, line in enumerate(lines):
                # Parse cells
                cells = [
                    self._clean_cell(cell)
                    for cell in line.strip("|").split("|")
                ]
                
                if j == 0:
                    headers = cells
                elif j == 1 and all(re.match(r'^[-:]+$', c.strip()) for c in cells):
                    # Skip separator line
                    continue
                else:
                    rows.append(cells)
            
            if headers or rows:
                table = self._create_table_content(
                    headers=headers,
                    rows=rows,
                    source_file=source_file,
                )
                table.markdown = table_text
                tables.append(table)
        
        return tables
    
    def extract_markdown_file(self, file_path: str) -> list[TableContent]:
        """Extract tables from a Markdown file."""
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        return self.extract_markdown(content, file_path)
    
    def extract_json(self, file_path: str) -> TableContent:
        """Extract table from JSON array."""
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        if not isinstance(data, list):
            if isinstance(data, dict) and any(isinstance(v, list) for v in data.values()):
                # Find the first list in the dict
                for key, value in data.items():
                    if isinstance(value, list):
                        data = value
                        break
            else:
                raise ValueError("JSON must contain an array")
        
        if not data:
            return self._create_table_content([], [], file_path)
        
        # Get headers from first object
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            rows = [[self._clean_cell(item.get(h)) for h in headers] for item in data]
        else:
            headers = [f"col_{i}" for i in range(len(data[0]))]
            rows = [[self._clean_cell(cell) for cell in row] for row in data]
        
        return self._create_table_content(headers, rows, file_path)
    
    def _clean_cell(self, value: Any) -> str:
        """Clean a cell value."""
        if value is None:
            return ""
        
        text = str(value)
        
        if self.config.clean_whitespace:
            text = " ".join(text.split())
        
        return text
    
    def _create_table_content(
        self,
        headers: list[str],
        rows: list[list[str]],
        source_file: str,
        caption: Optional[str] = None,
    ) -> TableContent:
        """Create a TableContent object."""
        # Ensure consistent column count
        if headers:
            num_cols = len(headers)
        elif rows:
            num_cols = max(len(row) for row in rows)
            headers = [f"col_{i}" for i in range(num_cols)]
        else:
            num_cols = 0
            headers = []
        
        # Pad rows to match header count
        padded_rows = []
        for row in rows:
            if len(row) < num_cols:
                row = row + [""] * (num_cols - len(row))
            elif len(row) > num_cols:
                row = row[:num_cols]
            padded_rows.append(row)
        
        # Infer types if enabled
        if self.config.infer_types:
            padded_rows = self._infer_and_convert_types(padded_rows)
        
        table = TableContent(
            id=generate_content_id(source_file, ContentType.TABLE, len(rows)),
            content_type=ContentType.TABLE,
            source_file=source_file,
            text="",  # Generated in __post_init__
            headers=headers,
            rows=padded_rows,
            caption=caption,
        )
        
        return table
    
    def _infer_and_convert_types(self, rows: list[list[str]]) -> list[list[Any]]:
        """Infer and convert cell types (numbers, etc.)."""
        if not rows:
            return rows
        
        converted = []
        
        for row in rows:
            new_row = []
            for cell in row:
                # Try to convert to number
                if isinstance(cell, str):
                    # Remove common formatting
                    clean = cell.replace(",", "").replace("$", "").replace("%", "").strip()
                    
                    try:
                        if "." in clean:
                            new_row.append(float(clean))
                        else:
                            new_row.append(int(clean))
                    except ValueError:
                        new_row.append(cell)
                else:
                    new_row.append(cell)
            
            converted.append(new_row)
        
        return converted
    
    def chunk_table(self, table: TableContent) -> list[TableContent]:
        """Split a large table into smaller chunks."""
        if len(table.rows) <= self.config.max_rows_per_chunk:
            return [table]
        
        chunks = []
        
        for i in range(0, len(table.rows), self.config.max_rows_per_chunk):
            chunk_rows = table.rows[i:i + self.config.max_rows_per_chunk]
            
            chunk = TableContent(
                id=f"{table.id}_chunk_{len(chunks)}",
                content_type=ContentType.TABLE,
                source_file=table.source_file,
                text="",
                headers=table.headers if self.config.include_headers_in_chunks else [],
                rows=chunk_rows,
                caption=f"{table.caption} (part {len(chunks) + 1})" if table.caption else None,
                metadata={
                    **table.metadata,
                    "chunk_index": len(chunks),
                    "start_row": i,
                    "end_row": i + len(chunk_rows),
                    "total_rows": len(table.rows),
                },
            )
            
            chunks.append(chunk)
        
        return chunks


class TableQueryEngine:
    """
    Enables structured queries on tables.
    
    Supports:
    - Column selection
    - Filtering
    - Aggregation
    - Sorting
    """
    
    def __init__(self, table: TableContent):
        self.table = table
        self._df = None
    
    @property
    def df(self):
        """Get pandas DataFrame (lazy loading)."""
        if self._df is None:
            try:
                import pandas as pd
                self._df = pd.DataFrame(self.table.rows, columns=self.table.headers)
            except ImportError:
                raise ImportError("Install pandas for table queries: pip install pandas")
        return self._df
    
    def select(self, columns: list[str]) -> list[dict]:
        """Select specific columns."""
        return self.df[columns].to_dict(orient="records")
    
    def filter(self, column: str, operator: str, value: Any) -> list[dict]:
        """Filter rows based on a condition."""
        df = self.df
        
        if operator == "==":
            mask = df[column] == value
        elif operator == "!=":
            mask = df[column] != value
        elif operator == ">":
            mask = df[column] > value
        elif operator == ">=":
            mask = df[column] >= value
        elif operator == "<":
            mask = df[column] < value
        elif operator == "<=":
            mask = df[column] <= value
        elif operator == "contains":
            mask = df[column].astype(str).str.contains(str(value), case=False)
        else:
            raise ValueError(f"Unknown operator: {operator}")
        
        return df[mask].to_dict(orient="records")
    
    def aggregate(self, column: str, operation: str) -> Any:
        """Aggregate a column."""
        col = self.df[column]
        
        if operation == "sum":
            return col.sum()
        elif operation == "mean":
            return col.mean()
        elif operation == "min":
            return col.min()
        elif operation == "max":
            return col.max()
        elif operation == "count":
            return len(col)
        elif operation == "unique":
            return col.nunique()
        else:
            raise ValueError(f"Unknown operation: {operation}")
    
    def sort(self, column: str, ascending: bool = True) -> list[dict]:
        """Sort by a column."""
        return self.df.sort_values(column, ascending=ascending).to_dict(orient="records")
    
    def head(self, n: int = 5) -> list[dict]:
        """Get first n rows."""
        return self.df.head(n).to_dict(orient="records")
    
    def describe(self) -> dict:
        """Get summary statistics."""
        return {
            "num_rows": len(self.table.rows),
            "num_cols": len(self.table.headers),
            "columns": self.table.headers,
            "sample": self.head(3),
        }


def extract_table(file_path: str) -> list[TableContent]:
    """Extract tables from a file."""
    extractor = TableExtractor()
    return extractor.extract_file(file_path)
